from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Sum, F
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
from decimal import Decimal
import json
import calendar
from openpyxl import load_workbook
from .models import (
    Student, Instructor, ClassSession, Attendance, Payment, 
    ClassType, InstructorAssignment, BankTransaction, PaymentAllocation,
    IncomeCategory, IncomeAllocation, ExpenseCategory, ExpenseAllocation,
    Seminar, SeminarPaymentAllocation, MembershipPaymentAllocation,
    MonthlyInstructorPayment, MonthlyFederationPayment, InstructorPaymentAllocation,
    PaymentCellComment
)
from .forms import BankTransactionUploadForm, PaymentAllocationForm


def login_view(request):
    """Нэвтрэх хуудас"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')  # Энэ нь имэйл байх болно
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            auth_login(request, user)
            messages.success(request, 'Амжилттай нэвтэрлээ!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Имэйл эсвэл нууц үг буруу байна.')
    
    return render(request, 'aikido_app/login.html')


def logout_view(request):
    """Гарах"""
    auth_logout(request)
    messages.success(request, 'Амжилттай гарлаа!')
    return redirect('login')


def register_view(request):
    """Бүртгүүлэх хуудас"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        # Validation
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Энэ нэвтрэх нэр аль хэдийн бүртгэгдсэн байна.')
        elif User.objects.filter(email=email).exists():
            messages.error(request, 'Энэ имэйл хаяг аль хэдийн бүртгэгдсэн байна.')
        elif password1 != password2:
            messages.error(request, 'Нууц үг таарахгүй байна.')
        elif len(password1) < 8:
            messages.error(request, 'Нууц үг хамгийн багадаа 8 тэмдэгт байх ёстой.')
        else:
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name
            )
            messages.success(request, 'Амжилттай бүртгэгдлээ! Нэвтрэж орно уу.')
            return redirect('login')
    
    return render(request, 'aikido_app/register.html')


@login_required
def dashboard(request):
    """Хяналтын самбар - зөвхөн админ"""
    # Redirect non-admin users to class schedule
    if not request.user.is_staff:
        return redirect('class_schedule')
    
    # Stats
    total_students = Student.objects.filter(is_active=True).count()
    total_instructors = Instructor.objects.filter(is_active=True).count()
    
    # This week's classes
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    classes_this_week = ClassSession.objects.filter(
        date__gte=week_start,
        date__lte=week_end
    ).count()
    
    # This month's revenue
    month_start = today.replace(day=1)
    payments_this_month = Payment.objects.filter(
        transaction_date__gte=month_start,
        is_verified=True
    )
    total_revenue = payments_this_month.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # New students this month
    new_students_this_month = Student.objects.filter(
        enrollment_date__gte=month_start
    ).count()
    
    # Attendance last week
    last_week_start = week_start - timedelta(days=7)
    attendance_last_week = []
    for i in range(7):
        day = last_week_start + timedelta(days=i)
        day_sessions = ClassSession.objects.filter(date=day)
        total_attendance = Attendance.objects.filter(session__date=day).count()
        present_count = Attendance.objects.filter(session__date=day, is_present=True).count()
        
        percentage = (present_count / total_attendance * 100) if total_attendance > 0 else 0
        
        attendance_last_week.append({
            'date': day,
            'weekday': day.strftime('%a'),
            'total': total_attendance,
            'present': present_count,
            'percentage': percentage
        })
    
    # Upcoming sessions
    upcoming_sessions = ClassSession.objects.filter(
        date__gte=today
    ).order_by('date', 'start_time')[:5]
    
    # Recent payments
    recent_payments = Payment.objects.select_related('student').order_by('-transaction_date')[:10]
    
    context = {
        'total_students': total_students,
        'total_instructors': total_instructors,
        'classes_this_week': classes_this_week,
        'total_revenue': total_revenue,
        'new_students_this_month': new_students_this_month,
        'payments_this_month': payments_this_month.count(),
        'attendance_last_week': attendance_last_week,
        'upcoming_sessions': upcoming_sessions,
        'recent_payments': recent_payments,
    }
    
    return render(request, 'aikido_app/dashboard.html', context)


@login_required
def student_list(request):
    """Сурагчдын жагсаалт"""
    # Get sort parameter from query string
    sort_by = request.GET.get('sort', '-enrollment_date')
    
    # Valid sort options
    valid_sorts = {
        'name_asc': 'first_name',
        'name_desc': '-first_name',
        'date_asc': 'enrollment_date',
        'date_desc': '-enrollment_date',
        'rank_asc': 'dan_rank',
        'rank_desc': '-dan_rank',
    }
    
    # Use the sort parameter if valid, otherwise default
    order_by = valid_sorts.get(sort_by, 'first_name')
    
    students = Student.objects.prefetch_related('class_types').all().order_by(order_by, 'last_name', 'first_name')
    
    return render(request, 'aikido_app/student_list.html', {
        'students': students,
        'current_sort': sort_by
    })


@login_required
def student_create(request):
    """Сурагч бүртгэх"""
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone = request.POST.get('phone')
        email = request.POST.get('email', '')
        date_of_birth = request.POST.get('date_of_birth') or None
        kyu_rank = request.POST.get('kyu_rank') or None
        dan_rank = request.POST.get('dan_rank') or None
        current_rank_date = request.POST.get('current_rank_date') or None
        monthly_fee = request.POST.get('monthly_fee') or None
        fee_note = request.POST.get('fee_note', '')
        class_types = request.POST.getlist('class_types')
        is_active = request.POST.get('is_active') == 'on'
        
        # Validation
        if not all([first_name, last_name, phone]):
            messages.error(request, 'Заавал бөглөх талбаруудыг бөглөнө үү.')
        else:
            try:
                student = Student.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone,
                    email=email,
                    date_of_birth=date_of_birth,
                    kyu_rank=int(kyu_rank) if kyu_rank else None,
                    dan_rank=int(dan_rank) if dan_rank else None,
                    current_rank_date=current_rank_date,
                    monthly_fee=monthly_fee,
                    fee_note=fee_note,
                    is_active=is_active
                )
                if class_types:
                    student.class_types.set(class_types)
                messages.success(request, f'{student} амжилттай бүртгэгдлээ!')
                return redirect('student_list')
            except Exception as e:
                messages.error(request, f'Алдаа гарлаа: {str(e)}')
    
    class_type_list = ClassType.objects.all()
    return render(request, 'aikido_app/student_form.html', {'class_types': class_type_list})


@login_required
def student_edit(request, pk):
    """Сурагчийн мэдээлэл засах"""
    try:
        student = Student.objects.get(pk=pk)
    except Student.DoesNotExist:
        messages.error(request, 'Сурагч олдсонгүй.')
        return redirect('student_list')
    
    if request.method == 'POST':
        student.first_name = request.POST.get('first_name')
        student.last_name = request.POST.get('last_name')
        student.phone = request.POST.get('phone')
        student.email = request.POST.get('email', '')
        date_of_birth = request.POST.get('date_of_birth')
        student.date_of_birth = date_of_birth if date_of_birth else None
        kyu_rank = request.POST.get('kyu_rank')
        dan_rank = request.POST.get('dan_rank')
        current_rank_date = request.POST.get('current_rank_date')
        monthly_fee = request.POST.get('monthly_fee')
        student.kyu_rank = int(kyu_rank) if kyu_rank else None
        student.dan_rank = int(dan_rank) if dan_rank else None
        student.current_rank_date = current_rank_date if current_rank_date else None
        student.monthly_fee = monthly_fee if monthly_fee else None
        student.fee_note = request.POST.get('fee_note', '')
        student.is_active = request.POST.get('is_active') == 'on'
        class_types = request.POST.getlist('class_types')
        
        try:
            student.save()
            if class_types:
                student.class_types.set(class_types)
            messages.success(request, f'{student} амжилттай шинэчлэгдлээ!')
            return redirect('student_list')
        except Exception as e:
            messages.error(request, f'Алдаа гарлаа: {str(e)}')
    
    class_type_list = ClassType.objects.all()
    return render(request, 'aikido_app/student_form.html', {
        'student': student, 
        'class_types': class_type_list
    })


@login_required
def student_delete(request, pk):
    """Сурагч устгах"""
    try:
        student = Student.objects.get(pk=pk)
        name = str(student)
        student.delete()
        messages.success(request, f'{name} амжилттай устгагдлаа!')
    except Student.DoesNotExist:
        messages.error(request, 'Сурагч олдсонгүй.')
    return redirect('student_list')


@login_required
def instructor_list(request):
    """Багш нарын жагсаалт"""
    # Check if user is an instructor (non-staff)
    is_instructor_user = hasattr(request.user, 'instructor_profile') and not request.user.is_staff
    
    # Get sort parameter from query string
    sort_by = request.GET.get('sort', 'name_asc')
    
    # Valid sort options
    valid_sorts = {
        'name_asc': 'last_name',
        'name_desc': '-last_name',
        'date_asc': 'hire_date',
        'date_desc': '-hire_date',
        'rank_asc': 'dan_rank',
        'rank_desc': '-dan_rank',
    }
    
    # Use the sort parameter if valid, otherwise default
    order_by = valid_sorts.get(sort_by, 'last_name')
    
    # Get filter parameter
    show_inactive = request.GET.get('show_inactive', 'false') == 'true'
    
    if is_instructor_user:
        # Instructor users only see themselves
        instructors = Instructor.objects.filter(user=request.user)
    else:
        # Admin/staff see all instructors
        if show_inactive:
            instructors = Instructor.objects.all()
        else:
            instructors = Instructor.objects.filter(is_active=True)
    
    instructors = instructors.order_by(order_by, 'last_name', 'first_name')
    
    # Calculate statistics for each instructor
    for instructor in instructors:
        # Count lead and assistant roles by class type
        lead_morning = InstructorAssignment.objects.filter(
            instructor=instructor,
            role='LEAD',
            session__class_type__name=ClassType.MORNING
        ).count()
        
        assistant_morning = InstructorAssignment.objects.filter(
            instructor=instructor,
            role='ASSISTANT',
            session__class_type__name=ClassType.MORNING
        ).count()
        
        lead_evening = InstructorAssignment.objects.filter(
            instructor=instructor,
            role='LEAD',
            session__class_type__name=ClassType.EVENING
        ).count()
        
        assistant_evening = InstructorAssignment.objects.filter(
            instructor=instructor,
            role='ASSISTANT',
            session__class_type__name=ClassType.EVENING
        ).count()
        
        lead_children = InstructorAssignment.objects.filter(
            instructor=instructor,
            role='LEAD',
            session__class_type__name=ClassType.CHILDREN
        ).count()
        
        assistant_children = InstructorAssignment.objects.filter(
            instructor=instructor,
            role='ASSISTANT',
            session__class_type__name=ClassType.CHILDREN
        ).count()
        
        instructor.stats = {
            'morning': {'lead': lead_morning, 'assistant': assistant_morning},
            'evening': {'lead': lead_evening, 'assistant': assistant_evening},
            'children': {'lead': lead_children, 'assistant': assistant_children},
            'total_lead': lead_morning + lead_evening + lead_children,
            'total_assistant': assistant_morning + assistant_evening + assistant_children,
        }
    
    return render(request, 'aikido_app/instructor_list.html', {
        'instructors': instructors,
        'current_sort': sort_by,
        'show_inactive': show_inactive
    })


@login_required
def instructor_create(request):
    """Багш бүртгэх"""
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone = request.POST.get('phone')
        email = request.POST.get('email', '')
        hire_date = request.POST.get('hire_date')
        kyu_rank = request.POST.get('kyu_rank') or None
        dan_rank = request.POST.get('dan_rank') or None
        current_rank_date = request.POST.get('current_rank_date') or None
        is_active = request.POST.get('is_active') == 'on'
        
        # Validation
        if not all([first_name, last_name, phone, hire_date]):
            messages.error(request, 'Заавал бөглөх талбаруудыг бөглөнө үү.')
        else:
            try:
                instructor = Instructor.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone,
                    email=email,
                    hire_date=hire_date,
                    kyu_rank=int(kyu_rank) if kyu_rank else None,
                    dan_rank=int(dan_rank) if dan_rank else None,
                    current_rank_date=current_rank_date if current_rank_date else None,
                    is_active=is_active
                )
                
                # Set allowed class types
                allowed_class_type_ids = request.POST.getlist('allowed_class_types')
                instructor.allowed_class_types.set(allowed_class_type_ids)
                
                messages.success(request, f'{instructor} амжилттай бүртгэгдлээ!')
                return redirect('instructor_list')
            except Exception as e:
                messages.error(request, f'Алдаа гарлаа: {str(e)}')
    
    context = {
        'class_types': ClassType.objects.all(),
    }
    return render(request, 'aikido_app/instructor_form.html', context)


@login_required
def instructor_edit(request, pk):
    """Багшийн мэдээлэл засах"""
    try:
        instructor = Instructor.objects.get(pk=pk)
    except Instructor.DoesNotExist:
        messages.error(request, 'Багш олдсонгүй.')
        return redirect('instructor_list')
    
    if request.method == 'POST':
        instructor.first_name = request.POST.get('first_name')
        instructor.last_name = request.POST.get('last_name')
        instructor.phone = request.POST.get('phone')
        instructor.email = request.POST.get('email', '')
        instructor.hire_date = request.POST.get('hire_date')
        kyu_rank = request.POST.get('kyu_rank')
        dan_rank = request.POST.get('dan_rank')
        current_rank_date = request.POST.get('current_rank_date')
        instructor.kyu_rank = int(kyu_rank) if kyu_rank else None
        instructor.dan_rank = int(dan_rank) if dan_rank else None
        instructor.current_rank_date = current_rank_date if current_rank_date else None
        instructor.is_active = request.POST.get('is_active') == 'on'
        
        try:
            instructor.save()
            
            # Update allowed class types
            allowed_class_type_ids = request.POST.getlist('allowed_class_types')
            instructor.allowed_class_types.set(allowed_class_type_ids)
            
            messages.success(request, f'{instructor} амжилттай шинэчлэгдлээ!')
            return redirect('instructor_list')
        except Exception as e:
            messages.error(request, f'Алдаа гарлаа: {str(e)}')
    
    context = {
        'instructor': instructor,
        'class_types': ClassType.objects.all(),
    }
    return render(request, 'aikido_app/instructor_form.html', context)


@login_required
def instructor_delete(request, pk):
    """Багш устгах"""
    try:
        instructor = Instructor.objects.get(pk=pk)
        name = str(instructor)
        instructor.delete()
        messages.success(request, f'{name} амжилттай устгагдлаа!')
    except Instructor.DoesNotExist:
        messages.error(request, 'Багш олдсонгүй.')
    except Exception as e:
        messages.error(request, f'Алдаа гарлаа: {str(e)}')
    
    return redirect('instructor_list')


@login_required
def class_schedule(request):
    """Хичээлийн хуваарь"""
    weekdays = ['Даваа', 'Мягмар', 'Лхагва', 'Пүрэв', 'Баасан', 'Бямба', 'Ням']
    return render(request, 'aikido_app/class_schedule.html', {'weekdays': weekdays})


@login_required
def attendance_list(request):
    """Ирцийн жагсаалт"""
    # Check if user is a student
    is_student_user = hasattr(request.user, 'student_profile') and request.user.student_profile
    
    if is_student_user:
        # Students see only their own attendance
        attendances = Attendance.objects.select_related(
            'student', 'session', 'recorded_by'
        ).filter(student=request.user.student_profile).order_by('-session__date')[:50]
    else:
        # Admins and instructors see all attendance
        attendances = Attendance.objects.select_related(
            'student', 'session', 'recorded_by'
        ).order_by('-session__date')[:50]
    
    context = {
        'attendances': attendances,
        'is_student_user': is_student_user
    }
    
    return render(request, 'aikido_app/attendance_list.html', context)


@login_required
def income_list(request):
    """Бусад орлогын жагсаалт - Семинар, гишүүнчлэл гэх мэт"""
    from django.db.models import Sum, Count, Q
    
    # Get filter parameters
    month_filter = request.GET.get('month', '')
    category_filter = request.GET.get('category', '')
    
    # Get income allocations (not student payments)
    allocations = IncomeAllocation.objects.select_related(
        'bank_transaction', 'income_category', 'created_by'
    ).order_by('-income_date', '-created_at')
    
    # Apply filters
    if month_filter:
        try:
            year, month = month_filter.split('-')
            allocations = allocations.filter(
                income_date__year=year,
                income_date__month=month
            )
        except:
            pass
    
    if category_filter:
        allocations = allocations.filter(income_category__id=category_filter)
    
    # Get unique months and categories for filters
    months = IncomeAllocation.objects.dates('income_date', 'month', order='DESC')
    categories = IncomeCategory.objects.all().order_by('name')
    
    # Calculate summary
    summary = allocations.aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id')
    )
    
    # Group by category for summary
    category_summary = allocations.values(
        'income_category__name'
    ).annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')
    
    context = {
        'allocations': allocations[:100],
        'months': months,
        'categories': categories,
        'selected_month': month_filter,
        'selected_category': category_filter,
        'summary': summary,
        'category_summary': category_summary,
    }
    return render(request, 'aikido_app/income_list.html', context)

@login_required
def seminar_payment_list(request):
    """Семинарын төлбөрийн жагсаалт"""
    from collections import defaultdict
    from django.db.models import Sum, Count, Avg
    
    # Get filters
    selected_seminar = request.GET.get('seminar', '')
    selected_student = request.GET.get('student', '')
    
    # Query payments
    payments = SeminarPaymentAllocation.objects.select_related(
        'student', 'seminar', 'bank_transaction', 'created_by'
    ).order_by('-seminar__seminar_date', 'student__last_name')
    
    if selected_seminar:
        payments = payments.filter(seminar_id=selected_seminar)
    
    if selected_student:
        payments = payments.filter(
            student__first_name__icontains=selected_student
        ) | payments.filter(
            student__last_name__icontains=selected_student
        )
    
    # Calculate summary
    summary = payments.aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id'),
        average_amount=Avg('amount')
    )
    
    # Group by seminar
    payments_by_seminar = defaultdict(list)
    for payment in payments:
        payments_by_seminar[payment.seminar].append(payment)
    
    # Get all seminars for filter
    seminars = Seminar.objects.filter(is_active=True).order_by('-seminar_date')
    
    context = {
        'payments_by_seminar': dict(payments_by_seminar),
        'seminars': seminars,
        'selected_seminar': selected_seminar,
        'selected_student': selected_student,
        'total_amount': summary['total_amount'] or 0,
        'total_count': summary['total_count'] or 0,
        'average_amount': summary['average_amount'] or 0,
    }
    
    return render(request, 'aikido_app/seminar_payment_list.html', context)

@login_required
def membership_payment_list(request):
    """Гишүүнчлэлийн төлбөрийн жагсаалт"""
    from collections import defaultdict
    from django.db.models import Sum, Count, Avg
    
    # Get filters
    selected_month = request.GET.get('month', '')
    selected_student = request.GET.get('student', '')
    
    # Query payments
    payments = MembershipPaymentAllocation.objects.select_related(
        'student', 'bank_transaction', 'created_by'
    ).order_by('-payment_month', 'student__last_name')
    
    if selected_month:
        from datetime import datetime
        month_date = datetime.strptime(selected_month, '%Y-%m').date()
        payments = payments.filter(payment_month=month_date)
    
    if selected_student:
        payments = payments.filter(
            student__first_name__icontains=selected_student
        ) | payments.filter(
            student__last_name__icontains=selected_student
        )
    
    # Calculate summary
    summary = payments.aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id'),
        average_amount=Avg('amount')
    )
    
    # Group by month
    payments_by_month = defaultdict(list)
    for payment in payments:
        payments_by_month[payment.payment_month].append(payment)
    
    context = {
        'payments_by_month': dict(payments_by_month),
        'selected_month': selected_month,
        'selected_student': selected_student,
        'total_amount': summary['total_amount'] or 0,
        'total_count': summary['total_count'] or 0,
        'average_amount': summary['average_amount'] or 0,
    }
    
    return render(request, 'aikido_app/membership_payment_list.html', context)

@login_required
def payment_list(request):
    """Төлбөрийн жагсаалт - Pivot хүснэгт хэлбэрээр"""
    from collections import defaultdict
    from datetime import datetime, date
    
    # Get selected year from request, default to current year
    selected_year = int(request.GET.get('year', datetime.now().year))
    
    # Handle POST request for updating cell comments
    if request.method == 'POST' and request.POST.get('action') == 'update_cell':
        student_id = request.POST.get('student_id')
        month_str = request.POST.get('month')
        comment = request.POST.get('comment', '')
        highlight_color = request.POST.get('highlight_color', '')
        
        try:
            student = Student.objects.get(id=student_id)
            month_date = datetime.strptime(month_str, '%Y-%m-%d').date()
            
            # Get or create cell comment
            cell_comment, created = PaymentCellComment.objects.get_or_create(
                student=student,
                month=month_date,
                defaults={'comment': comment, 'highlight_color': highlight_color}
            )
            
            if not created:
                cell_comment.comment = comment
                cell_comment.highlight_color = highlight_color
                cell_comment.save()
            
            messages.success(request, 'Мэдээлэл шинэчлэгдлээ!')
        except (Student.DoesNotExist, ValueError) as e:
            messages.error(request, f'Алдаа гарлаа: {str(e)}')
        
        from django.urls import reverse
        redirect_url = reverse('payment_list') + f'?year={selected_year}'
        return redirect(redirect_url)
    
    # Filter allocations by selected year
    allocations = PaymentAllocation.objects.filter(
        payment_month__year=selected_year
    ).select_related(
        'bank_transaction', 'student'
    ).prefetch_related('student__class_types').order_by(
        'student__class_types__name', 'student__first_name', 'student__last_name',  'payment_month'
    )
    
    # Get all students (both with and without payments)
    all_students = Student.objects.prefetch_related('class_types').distinct().order_by('first_name', 'last_name')
    
    # Get attendance data grouped by student and month for selected year
    from django.db.models import Count, Q
    attendance_data = defaultdict(dict)
    
    # Get attendance records for selected year
    attendances = Attendance.objects.filter(
        is_present=True,
        session__date__year=selected_year
    ).select_related('session', 'student').order_by('session__date')
    
    for attendance in attendances:
        student_id = attendance.student.id
        # Get year-month from session date
        session_date = attendance.session.date
        month_key = date(session_date.year, session_date.month, 1)
        
        if month_key not in attendance_data[student_id]:
            attendance_data[student_id][month_key] = 0
        attendance_data[student_id][month_key] += 1
    
    # Create fixed months for selected year (January to December)
    months = [date(selected_year, month, 1) for month in range(1, 13)]
    
    # Get all cell comments for selected year
    cell_comments = {}
    for comment in PaymentCellComment.objects.filter(
        month__year=selected_year
    ).select_related('student').all():
        key = (comment.student.id, comment.month)
        cell_comments[key] = {
            'comment': comment.comment,
            'highlight_color': comment.highlight_color
        }
    
    # Build pivot data structure
    # pivot_data[student_id][month_date] = {'amount': X, 'date': Y, 'allocations': [], 'attendance': N}
    pivot_data = defaultdict(dict)
    
    for allocation in allocations:
        student_id = allocation.student.id
        month_key = allocation.payment_month
        
        if month_key not in pivot_data[student_id]:
            pivot_data[student_id][month_key] = {
                'amount': 0,
                'date': allocation.bank_transaction.transaction_date if allocation.bank_transaction else None,
                'allocations': [],
                'total_attendance': 0,
                'actual_attendance': attendance_data.get(student_id, {}).get(month_key, 0),
                'comments': [],
                'highlight_color': ''
            }
        
        pivot_data[student_id][month_key]['amount'] += allocation.amount
        pivot_data[student_id][month_key]['total_attendance'] += allocation.attendance_count
        if allocation.comment:
            pivot_data[student_id][month_key]['comments'].append(allocation.comment)
        if allocation.highlight_color and not pivot_data[student_id][month_key]['highlight_color']:
            pivot_data[student_id][month_key]['highlight_color'] = allocation.highlight_color
        
        pivot_data[student_id][month_key]['allocations'].append({
            'id': allocation.id,
            'amount': float(allocation.amount),
            'date': allocation.bank_transaction.transaction_date.strftime('%Y-%m-%d') if allocation.bank_transaction and allocation.bank_transaction.transaction_date else None,
            'attendance_count': allocation.attendance_count,
            'comment': allocation.comment or '',
            'highlight_color': allocation.highlight_color or ''
        })
    
    # Add attendance data for months without payments
    for student_id, month_attendance in attendance_data.items():
        for month_key, attendance_count in month_attendance.items():
            if month_key not in pivot_data[student_id]:
                pivot_data[student_id][month_key] = {
                    'amount': 0,
                    'date': None,
                    'allocations': [],
                    'total_attendance': 0,
                    'actual_attendance': attendance_count,
                    'comments': [],
                    'highlight_color': ''
                }
            else:
                # Update actual attendance if not set
                if 'actual_attendance' not in pivot_data[student_id][month_key]:
                    pivot_data[student_id][month_key]['actual_attendance'] = attendance_count
    
    # Group students by class type
    class_groups = defaultdict(list)
    for student in all_students:
        student_classes = student.class_types.all()
        if student_classes:
            for class_type in student_classes:
                class_groups[class_type.name].append(student)
        else:
            class_groups['Ангигүй'].append(student)
    
    # Prepare rows grouped by class type with row totals
    groups = []
    for class_name in sorted(class_groups.keys()):
        rows = []
        group_total = 0
        group_column_totals = []
        
        for student in class_groups[class_name]:
            row = {
                'student': student,
                'months': [],
                'row_total': 0
            }
            for month in months:
                # Get cell comment for this student-month combination
                cell_comment_data = cell_comments.get((student.id, month), {})
                
                if month in pivot_data[student.id]:
                    month_data = pivot_data[student.id][month].copy()
                    # Apply cell comment if exists (overrides allocation-based color/comment)
                    if cell_comment_data:
                        if cell_comment_data.get('highlight_color'):
                            month_data['highlight_color'] = cell_comment_data['highlight_color']
                        if cell_comment_data.get('comment'):
                            if month_data['comments']:
                                month_data['comments'].append(cell_comment_data['comment'])
                            else:
                                month_data['comments'] = [cell_comment_data['comment']]
                    # Convert allocations to JSON string for template
                    month_data['allocations_json'] = json.dumps(month_data['allocations'])
                    month_data['month_str'] = month.strftime('%Y-%m-%d')
                    month_data['has_data'] = True
                    row['months'].append(month_data)
                    row['row_total'] += month_data['amount']
                else:
                    # No payment/attendance, but may have cell comment
                    month_data = {
                        'amount': 0,
                        'date': None,
                        'allocations': [],
                        'allocations_json': '[]',
                        'total_attendance': 0,
                        'actual_attendance': 0,
                        'comments': [cell_comment_data.get('comment')] if cell_comment_data.get('comment') else [],
                        'highlight_color': cell_comment_data.get('highlight_color', ''),
                        'month_str': month.strftime('%Y-%m-%d'),
                        'has_data': bool(cell_comment_data)
                    }
                    row['months'].append(month_data)
            rows.append(row)
            group_total += row['row_total']
        
        # Calculate column totals for this group
        for month in months:
            month_total = sum(
                pivot_data[student.id].get(month, {}).get('amount', 0)
                for student in class_groups[class_name]
            )
            group_column_totals.append(month_total)
        
        groups.append({
            'class_name': class_name,
            'rows': rows,
            'group_total': group_total,
            'group_column_totals': group_column_totals
        })
    
    # Calculate column totals for each month
    column_totals = []
    for month in months:
        month_total = sum(
            pivot_data[student.id].get(month, {}).get('amount', 0)
            for student in all_students
        )
        column_totals.append(month_total)
    
    # Calculate summary for selected year
    from django.db.models import Sum, Count
    summary = PaymentAllocation.objects.filter(
        payment_month__year=selected_year
    ).aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id')
    )
    
    # Get available years for dropdown
    available_years = PaymentAllocation.objects.dates('payment_month', 'year', order='DESC')
    years_list = sorted(set([d.year for d in available_years]), reverse=True)
    if selected_year not in years_list and allocations.exists():
        years_list.append(selected_year)
        years_list.sort(reverse=True)
    
    # Add current year if not in list
    current_year = datetime.now().year
    if current_year not in years_list:
        years_list.insert(0, current_year)
    
    context = {
        'groups': groups,
        'months': months,
        'column_totals': column_totals,
        'summary': summary,
        'selected_year': selected_year,
        'years_list': years_list,
    }
    return render(request, 'aikido_app/payment_list.html', context)


@login_required
def attendance_record(request):
    """Ирц бүртгэх - Багшийн хуудас"""
    # Get current user's instructor record if exists
    try:
        instructor = Instructor.objects.get(user=request.user)
    except Instructor.DoesNotExist:
        # Check by email
        try:
            instructor = Instructor.objects.get(email=request.user.email)
        except Instructor.DoesNotExist:
            instructor = None
    
    # Handle AJAX POST request for saving attendance
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            attendance_data = data.get('attendance', [])
            instructor_assignments = data.get('instructor_assignments', [])
            class_type_name = data.get('class_type', '')
            
            # Save instructor assignments first
            for assignment_data in instructor_assignments:
                date_str = assignment_data.get('date')
                lead_instructor_id = assignment_data.get('lead_instructor_id')
                assistant_instructor_id = assignment_data.get('assistant_instructor_id')
                
                # Handle "NO_CLASS" marker - save empty session to indicate class was cancelled
                if lead_instructor_id == 'NO_CLASS' or assistant_instructor_id == 'NO_CLASS':
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                        
                        # Get or create class session
                        class_type = ClassType.objects.filter(name=class_type_name).first() or ClassType.objects.first()
                        session, created = ClassSession.objects.get_or_create(
                            date=date_obj,
                            class_type=class_type,
                            defaults={
                                'weekday': date_obj.weekday(),
                                'start_time': '09:00',
                                'end_time': '11:00',
                            }
                        )
                        
                        # Clear all existing assignments - this marks it as "no class"
                        InstructorAssignment.objects.filter(session=session).delete()
                    except Exception as e:
                        print(f"Error saving NO_CLASS marker: {e}")
                    continue
                
                # Skip if both are empty (and not NO_CLASS)
                if not lead_instructor_id and not assistant_instructor_id:
                    continue
                
                try:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                    
                    # Get or create class session
                    class_type = ClassType.objects.filter(name=class_type_name).first() or ClassType.objects.first()
                    session, created = ClassSession.objects.get_or_create(
                        date=date_obj,
                        class_type=class_type,
                        defaults={
                            'weekday': date_obj.weekday(),
                            'start_time': '09:00',
                            'end_time': '11:00',
                        }
                    )
                    
                    # Clear existing assignments for this session
                    InstructorAssignment.objects.filter(session=session).delete()
                    
                    # Create new assignments only if instructor is selected
                    if lead_instructor_id and lead_instructor_id.strip():
                        try:
                            lead_instructor = Instructor.objects.get(id=int(lead_instructor_id))
                            InstructorAssignment.objects.create(
                                session=session,
                                instructor=lead_instructor,
                                role='LEAD'
                            )
                        except (Instructor.DoesNotExist, ValueError):
                            pass
                    
                    if assistant_instructor_id and assistant_instructor_id.strip():
                        try:
                            assistant_instructor = Instructor.objects.get(id=int(assistant_instructor_id))
                            InstructorAssignment.objects.create(
                                session=session,
                                instructor=assistant_instructor,
                                role='ASSISTANT'
                            )
                        except (Instructor.DoesNotExist, ValueError):
                            pass
                except Exception as e:
                    print(f"Error saving instructor assignment: {e}")
            
            # Save attendance
            for item in attendance_data:
                student_id = item.get('student_id')
                date_str = item.get('date')
                
                try:
                    student = Student.objects.get(id=student_id)
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                    
                    # Get class type from request data
                    class_type = ClassType.objects.filter(name=class_type_name).first()
                    if not class_type:
                        class_type = student.class_types.first() or ClassType.objects.first()
                    
                    # Find or create session for this date and class type
                    session, created = ClassSession.objects.get_or_create(
                        date=date_obj,
                        class_type=class_type,
                        defaults={
                            'weekday': date_obj.weekday(),
                            'start_time': '09:00',
                            'end_time': '11:00',
                        }
                    )
                    
                    # Get is_present from request data (defaults to True if not specified)
                    is_present = item.get('is_present', True)
                    
                    if is_present:
                        # Create or update attendance
                        Attendance.objects.update_or_create(
                            session=session,
                            student=student,
                            defaults={
                                'is_present': True,
                                'recorded_by': instructor
                            }
                        )
                    else:
                        # Delete attendance if unchecked
                        Attendance.objects.filter(
                            session=session,
                            student=student
                        ).delete()
                except Student.DoesNotExist:
                    pass
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET request - display attendance form
    selected_class_type = request.GET.get('class_type', '')
    selected_month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    
    # Parse selected month
    year, month = map(int, selected_month.split('-'))
    
    # Get all class types
    class_types = ClassType.objects.all()
    
    # Filter students by class type
    students = Student.objects.filter(is_active=True).prefetch_related('class_types')
    
    if selected_class_type:
        students = students.filter(class_types__name=selected_class_type)
    
    students = students.order_by('first_name', 'last_name')
    
    # Generate dates based on class type
    dates = []
    
    if selected_class_type:
        # Get the ClassType object
        try:
            class_type_obj = ClassType.objects.get(name=selected_class_type)
            
            # Determine which weekdays to show
            if class_type_obj.name in [ClassType.MORNING, ClassType.EVENING]:
                # Mon, Wed, Fri (0, 2, 4)
                target_weekdays = [0, 2, 4]
                weekday_names = ['Даваа', 'Лхагва', 'Баасан']
            else:  # Children
                # Sat, Sun (5, 6)
                target_weekdays = [5, 6]
                weekday_names = ['Бямба', 'Ням']
        except ClassType.DoesNotExist:
            target_weekdays = [0, 2, 4, 5, 6]
            weekday_names = ['Даваа', 'Лхагва', 'Баасан', 'Бямба', 'Ням']
    else:
        # Show all days: Mon, Wed, Fri, Sat, Sun
        target_weekdays = [0, 2, 4, 5, 6]
        weekday_names = ['Даваа', 'Лхагва', 'Баасан', 'Бямба', 'Ням']
    
    # Get all days in the month
    num_days = calendar.monthrange(year, month)[1]
    
    for day in range(1, num_days + 1):
        date = datetime(year, month, day).date()
        if date.weekday() in target_weekdays:
            weekday_index = target_weekdays.index(date.weekday())
            dates.append({
                'date': date,
                'weekday': weekday_names[weekday_index]
            })
    
    # Get existing attendance for these students and dates
    if students.exists() and dates:
        date_list = [d['date'] for d in dates]
        attendances = Attendance.objects.filter(
            student__in=students,
            session__date__in=date_list,
            is_present=True
        ).select_related('student', 'session')
        
        # Build attendance map
        attendance_map = {}
        for att in attendances:
            if att.student.id not in attendance_map:
                attendance_map[att.student.id] = set()
            attendance_map[att.student.id].add(att.session.date.strftime('%Y-%m-%d'))
        
        # Add attendance info to students
        for student in students:
            student.attendance_dates = attendance_map.get(student.id, set())
            student.attendance_count = len(student.attendance_dates)
    
    # Get all active instructors for assignment
    all_instructors = Instructor.objects.filter(is_active=True).order_by('last_name', 'first_name')
    
    # Determine default instructors based on class type
    # Using actual instructor names from database:
    # ID 3: Галбадрах (Хүүхэд)
    # ID 1: Амгаланбаяр (Өглөө)
    # ID 2: Баясгалан (Орой)
    default_lead = None
    default_assistant = None
    
    try:
        if selected_class_type == ClassType.CHILDREN:
            # Хүүхдийн анги - Галбадрах
            default_instructor = Instructor.objects.get(id=3)  # Галбадрах
            default_lead = default_instructor
            default_assistant = default_instructor
        elif selected_class_type == ClassType.MORNING:
            # Өглөөний анги - Амгаланбаяр
            default_instructor = Instructor.objects.get(id=1)  # Амгаланбаяр
            default_lead = default_instructor
            default_assistant = default_instructor
        elif selected_class_type == ClassType.EVENING:
            # Оройны анги - Баясгалан
            default_instructor = Instructor.objects.get(id=2)  # Баясгалан
            default_lead = default_instructor
            default_assistant = default_instructor
    except Instructor.DoesNotExist:
        pass
    
    # Get existing instructor assignments for the dates
    if dates:
        date_list = [d['date'] for d in dates]
        
        # Filter sessions by date and optionally by class_type
        sessions_query = ClassSession.objects.filter(date__in=date_list)
        if selected_class_type:
            sessions_query = sessions_query.filter(class_type__name=selected_class_type)
        
        sessions_with_instructors = sessions_query.prefetch_related('instructor_assignments__instructor')
        
        # Build instructor assignment map: {date: {'lead': instructor, 'assistant': instructor, 'class_type': name, 'has_assignments': bool, 'is_cancelled': bool}}
        instructor_assignment_map = {}
        for session in sessions_with_instructors:
            date_str = session.date.strftime('%Y-%m-%d')
            has_assignments = session.instructor_assignments.exists()
            
            instructor_assignment_map[date_str] = {
                'class_type': session.class_type.name,
                'has_assignments': has_assignments,
                'is_cancelled': session.is_cancelled
            }
            
            for assignment in session.instructor_assignments.all():
                if assignment.role == 'LEAD':
                    instructor_assignment_map[date_str]['lead'] = assignment.instructor
                elif assignment.role == 'ASSISTANT':
                    instructor_assignment_map[date_str]['assistant'] = assignment.instructor
        
        # Add instructor info to dates
        for date_item in dates:
            date_str = date_item['date'].strftime('%Y-%m-%d')
            assignments = instructor_assignment_map.get(date_str, {})
            session_class_type = assignments.get('class_type')
            has_assignments = assignments.get('has_assignments', False)
            is_cancelled = assignments.get('is_cancelled', False)
            
            # Decision logic for showing instructors:
            if selected_class_type:
                # A specific class type is selected
                if not session_class_type:
                    # No session exists - show defaults
                    date_item['lead_instructor'] = default_lead
                    date_item['assistant_instructor'] = default_assistant
                elif session_class_type == selected_class_type:
                    # Session exists and matches selected class type
                    if is_cancelled:
                        # Session is cancelled - show nothing (empty)
                        date_item['lead_instructor'] = None
                        date_item['assistant_instructor'] = None
                    elif has_assignments:
                        # Has instructor assignments - show them
                        date_item['lead_instructor'] = assignments.get('lead')
                        date_item['assistant_instructor'] = assignments.get('assistant')
                    else:
                        # Session exists but no instructors assigned yet - show defaults
                        date_item['lead_instructor'] = default_lead
                        date_item['assistant_instructor'] = default_assistant
                else:
                    # Session exists but different class type - show nothing (empty)
                    date_item['lead_instructor'] = None
                    date_item['assistant_instructor'] = None
            else:
                # No class type selected (viewing all classes)
                # Only show assigned instructors, no defaults (unless cancelled)
                if is_cancelled:
                    date_item['lead_instructor'] = None
                    date_item['assistant_instructor'] = None
                else:
                    date_item['lead_instructor'] = assignments.get('lead')
                    date_item['assistant_instructor'] = assignments.get('assistant')
    
    # Calculate instructor statistics for the selected month and class type
    instructor_stats = []
    if dates:  # Changed: removed selected_class_type requirement
        date_list = [d['date'] for d in dates]
        
        for instructor in all_instructors:
            # Calculate stats by class type
            stats_by_class = {}
            total_lead = 0
            total_assistant = 0
            
            # Get stats for each class type
            for class_type in class_types:
                lead_count = InstructorAssignment.objects.filter(
                    instructor=instructor,
                    role='LEAD',
                    session__date__in=date_list,
                    session__class_type__name=class_type.name
                ).count()
                
                assistant_count = InstructorAssignment.objects.filter(
                    instructor=instructor,
                    role='ASSISTANT',
                    session__date__in=date_list,
                    session__class_type__name=class_type.name
                ).count()
                
                if lead_count > 0 or assistant_count > 0:
                    stats_by_class[class_type.name] = {
                        'class_type': class_type,
                        'lead': lead_count,
                        'assistant': assistant_count,
                        'total': lead_count + assistant_count
                    }
                
                total_lead += lead_count
                total_assistant += assistant_count
            
            total = total_lead + total_assistant
            
            # Always show instructors if they have any stats (regardless of selected_class_type)
            if total > 0:
                # If a specific class type is selected, use that class stats for main counts
                if selected_class_type and selected_class_type in stats_by_class:
                    instructor_stats.append({
                        'instructor': instructor,
                        'stats_by_class': stats_by_class,
                        'lead_count': stats_by_class[selected_class_type]['lead'],
                        'assistant_count': stats_by_class[selected_class_type]['assistant'],
                        'total': stats_by_class[selected_class_type]['total']
                    })
                else:
                    # No specific class selected or instructor doesn't have stats for selected class
                    # Show total across all classes
                    instructor_stats.append({
                        'instructor': instructor,
                        'stats_by_class': stats_by_class,
                        'lead_count': total_lead,
                        'assistant_count': total_assistant,
                        'total': total
                    })
        
        # Sort by total descending
        instructor_stats.sort(key=lambda x: x['total'], reverse=True)
    
    context = {
        'class_types': class_types,
        'selected_class_type': selected_class_type,
        'selected_month': selected_month,
        'students': students,
        'dates': dates,
        'all_instructors': all_instructors,
        'default_lead': default_lead,
        'default_assistant': default_assistant,
        'instructor_stats': instructor_stats,
    }
    
    return render(request, 'aikido_app/attendance_record.html', context)


@csrf_exempt
@login_required
def assign_instructors(request):
    """Тухайн өдрийн ахлах болон туслах багшийг томилох"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            date_str = data.get('date')
            lead_instructor_id = data.get('lead_instructor_id')
            assistant_instructor_id = data.get('assistant_instructor_id')
            class_type_name = data.get('class_type')
            
            # Validation
            if not date_str:
                return JsonResponse({
                    'success': False,
                    'error': 'Огноо дутуу байна'
                }, status=400)
            
            # Parse date
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Огнооны формат буруу: {str(e)}'
                }, status=400)
            
            # Get or create class session for this date
            class_type = ClassType.objects.filter(name=class_type_name).first() or ClassType.objects.first()
            
            if not class_type:
                return JsonResponse({
                    'success': False,
                    'error': 'Ангийн төрөл олдсонгүй'
                }, status=400)
            
            session, created = ClassSession.objects.get_or_create(
                date=date_obj,
                class_type=class_type,
                defaults={
                    'weekday': date_obj.weekday(),
                    'start_time': '19:30',
                    'end_time': '21:00',
                }
            )
            
            # Clear existing instructor assignments for this session
            InstructorAssignment.objects.filter(session=session).delete()
            
            # Handle NO_CLASS case - if either instructor is NO_CLASS, don't assign anyone
            if lead_instructor_id == 'NO_CLASS' or assistant_instructor_id == 'NO_CLASS':
                # Mark session as cancelled
                session.is_cancelled = True
                session.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Хичээл ороогүй гэж тэмдэглэгдлээ'
                })
            
            # Assign lead instructor
            if lead_instructor_id:
                try:
                    lead_instructor = Instructor.objects.get(id=lead_instructor_id)
                    InstructorAssignment.objects.create(
                        session=session,
                        instructor=lead_instructor,
                        role='LEAD'
                    )
                except Instructor.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': f'Ахлах багш (ID: {lead_instructor_id}) олдсонгүй'
                    }, status=400)
            
            # Assign assistant instructor
            if assistant_instructor_id:
                try:
                    assistant_instructor = Instructor.objects.get(id=assistant_instructor_id)
                    InstructorAssignment.objects.create(
                        session=session,
                        instructor=assistant_instructor,
                        role='ASSISTANT'
                    )
                except Instructor.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': f'Туслах багш (ID: {assistant_instructor_id}) олдсонгүй'
                    }, status=400)
            
            # If instructors were assigned, mark session as not cancelled
            if lead_instructor_id or assistant_instructor_id:
                session.is_cancelled = False
                session.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Багш амжилттай томилогдлоо'
            })
            
        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False,
                'error': f'JSON алдаа: {str(e)}'
            }, status=400)
        except Exception as e:
            # Log the full error for debugging
            import traceback
            print(f"Error in assign_instructors: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'error': f'Алдаа гарлаа: {str(e)}'
            }, status=400)


@login_required
def bank_transaction_list(request):
    """Банкны гүйлгээний жагсаалт"""
    from django.core.paginator import Paginator
    
    transactions = BankTransaction.objects.all().select_related().prefetch_related(
        'allocations__student', 'expense_allocations__expense_category'
    )
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        transactions = transactions.filter(status=status_filter)
    
    # Filter by month
    month_filter = request.GET.get('month')
    if month_filter:
        try:
            year, month = month_filter.split('-')
            transactions = transactions.filter(
                transaction_date__year=year,
                transaction_date__month=month
            )
        except:
            pass
    
    # Filter by transaction type (income/expense)
    type_filter = request.GET.get('type')
    if type_filter == 'income':
        transactions = transactions.filter(credit_amount__gt=0)
    elif type_filter == 'expense':
        transactions = transactions.exclude(debit_amount=0).exclude(debit_amount__isnull=True)
    
    # Order by date descending
    transactions = transactions.order_by('-transaction_date', '-imported_at')
    
    # Pagination
    paginator = Paginator(transactions, 50)  # 50 transactions per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get unique months for filter
    months = BankTransaction.objects.dates('transaction_date', 'month', order='DESC')
    
    # Calculate summary for current filter
    from django.db.models import Sum, Count, Q
    summary = transactions.aggregate(
        total_count=Count('id'),
        income_count=Count('id', filter=Q(credit_amount__gt=0)),
        expense_count=Count('id', filter=~Q(debit_amount=0) & ~Q(debit_amount__isnull=True)),
        total_income=Sum('credit_amount', filter=Q(credit_amount__gt=0)),
        total_expense=Sum('debit_amount', filter=~Q(debit_amount=0) & ~Q(debit_amount__isnull=True)),
    )
    
    context = {
        'page_obj': page_obj,
        'transactions': page_obj.object_list,
        'status_choices': BankTransaction.STATUS_CHOICES,
        'selected_status': status_filter,
        'selected_month': month_filter,
        'selected_type': type_filter,
        'months': months,
        'summary': summary,
    }
    return render(request, 'aikido_app/bank_transaction_list.html', context)


@login_required
def bank_transaction_upload(request):
    """Excel файл импортлох"""
    if request.method == 'POST':
        form = BankTransactionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                bank_format = form.cleaned_data['bank_format']
                start_row = form.cleaned_data['start_row']
                
                # Load Excel file
                workbook = load_workbook(excel_file, data_only=True)
                sheet = workbook.active
                
                # Determine column mapping
                if bank_format == 'standard':
                    # Find columns by header names
                    date_header = form.cleaned_data['date_header']
                    amount_header = form.cleaned_data['amount_header']
                    desc_header = form.cleaned_data['description_header']
                    payer_header = form.cleaned_data.get('payer_header', '')
                    
                    # Find column indices from first row
                    header_row = sheet[1]
                    date_col = None
                    opening_col = None
                    debit_col = None
                    credit_col = None
                    closing_col = None
                    desc_col = None
                    counterparty_col = None
                    
                    for idx, cell in enumerate(header_row, start=1):
                        if cell.value:
                            cell_value = str(cell.value).strip()
                            print(f"DEBUG: Column {cell.column_letter} = '{cell_value}'")
                            
                            if 'огноо' in cell_value.lower():
                                date_col = cell.column_letter
                            elif 'эхний' in cell_value.lower() and 'үлдэгдэл' in cell_value.lower():
                                opening_col = cell.column_letter
                            elif 'дебит' in cell_value.lower():
                                debit_col = cell.column_letter
                            elif 'кредит' in cell_value.lower():
                                credit_col = cell.column_letter
                            elif 'эцсийн' in cell_value.lower() and 'үлдэгдэл' in cell_value.lower():
                                closing_col = cell.column_letter
                            elif 'утга' in cell_value.lower():
                                desc_col = cell.column_letter
                            elif 'харьцсан' in cell_value.lower() or 'данс' in cell_value.lower():
                                counterparty_col = cell.column_letter
                    
                    print(f"DEBUG: Found columns - Date:{date_col}, Opening:{opening_col}, Debit:{debit_col}, Credit:{credit_col}, Closing:{closing_col}, Desc:{desc_col}, Counterparty:{counterparty_col}")
                    
                    if not date_col or not credit_col or not desc_col:
                        messages.error(request, f'Шаардлагатай баганууд олдсонгүй. Date:{date_col}, Credit:{credit_col}, Desc:{desc_col}')
                        return render(request, 'aikido_app/bank_transaction_upload.html', {'form': form})
                else:
                    # Custom format - use manual column letters
                    date_col = form.cleaned_data['date_column'].upper()
                    credit_col = form.cleaned_data['amount_column'].upper()
                    desc_col = form.cleaned_data['description_column'].upper()
                    counterparty_col = form.cleaned_data.get('payer_name_column', '').upper()
                    opening_col = None
                    debit_col = None
                    closing_col = None
                
                imported_count = 0
                errors = []
                skipped_count = 0
                
                print(f"DEBUG: Starting import from row {start_row} to {sheet.max_row}")
                print(f"DEBUG: Columns - Date:{date_col}, Credit:{credit_col}, Desc:{desc_col}")
                
                for row_num in range(start_row, sheet.max_row + 1):
                    try:
                        # Read all values
                        date_val = sheet[f'{date_col}{row_num}'].value
                        opening_val = sheet[f'{opening_col}{row_num}'].value if opening_col else None
                        debit_val = sheet[f'{debit_col}{row_num}'].value if debit_col else None
                        credit_val = sheet[f'{credit_col}{row_num}'].value if credit_col else None
                        closing_val = sheet[f'{closing_col}{row_num}'].value if closing_col else None
                        desc_val = sheet[f'{desc_col}{row_num}'].value
                        counterparty_val = sheet[f'{counterparty_col}{row_num}'].value if counterparty_col else ''
                        
                        print(f"DEBUG Row {row_num}: date={date_val}, opening={opening_val}, debit={debit_val}, credit={credit_val}, closing={closing_val}, desc={desc_val}")
                        
                        # Skip empty rows (only if date is missing)
                        if not date_val:
                            skipped_count += 1
                            continue
                        
                        # Parse all amounts
                        opening_balance = None
                        debit_amount = None
                        credit_amount = None
                        closing_balance = None
                        
                        if opening_val:
                            try:
                                opening_balance = Decimal(str(opening_val).replace(',', '').replace(' ', ''))
                            except:
                                pass
                        
                        if debit_val:
                            try:
                                debit_amount = Decimal(str(debit_val).replace(',', '').replace(' ', ''))
                            except:
                                pass
                        
                        if credit_val:
                            try:
                                credit_amount = Decimal(str(credit_val).replace(',', '').replace(' ', ''))
                            except:
                                pass
                        
                        if closing_val:
                            try:
                                closing_balance = Decimal(str(closing_val).replace(',', '').replace(' ', ''))
                            except:
                                pass
                        
                        # Determine main amount (use credit if available, otherwise debit)
                        main_amount = Decimal('0.00')
                        if credit_amount and credit_amount > 0:
                            main_amount = credit_amount
                        elif debit_amount and debit_amount != 0:
                            main_amount = abs(debit_amount)  # Use absolute value since debit is negative
                        
                        # Parse date
                        if isinstance(date_val, datetime):
                            transaction_date = date_val.date()
                        elif isinstance(date_val, str):
                            # Try multiple date formats (including datetime formats)
                            transaction_date = None
                            date_formats = [
                                '%Y/%m/%d %H:%M',    # 2025/07/30 12:15
                                '%Y/%m/%d  %H:%M',   # 2025/07/30  12:15 (double space)
                                '%Y-%m-%d %H:%M',
                                '%Y/%m/%d',
                                '%Y-%m-%d',
                                '%d.%m.%Y',
                                '%d/%m/%Y'
                            ]
                            for date_format in date_formats:
                                try:
                                    transaction_date = datetime.strptime(date_val.strip(), date_format).date()
                                    break
                                except:
                                    continue
                            
                            if not transaction_date:
                                print(f"DEBUG: Failed to parse date '{date_val}' on row {row_num}")
                                errors.append(f'Мөр {row_num}: Огнооны формат буруу - {date_val}')
                                error_count += 1
                                continue
                        else:
                            print(f"DEBUG: Invalid date type '{type(date_val)}' on row {row_num}")
                            errors.append(f'Мөр {row_num}: Огнооны формат буруу')
                            error_count += 1
                            continue
                        
                        # Check for duplicates before creating
                        # Check based on date, description, and either credit or debit amount
                        duplicate_filter = {
                            'transaction_date': transaction_date,
                            'description': str(desc_val)[:500] if desc_val else ''
                        }
                        
                        if credit_amount and credit_amount > 0:
                            duplicate_filter['credit_amount'] = credit_amount
                        elif debit_amount and debit_amount != 0:
                            duplicate_filter['debit_amount'] = debit_amount
                        
                        existing = BankTransaction.objects.filter(**duplicate_filter).exists()
                        
                        if existing:
                            skipped_count += 1
                            continue
                        
                        # Create transaction
                        BankTransaction.objects.create(
                            transaction_date=transaction_date,
                            opening_balance=opening_balance,
                            debit_amount=debit_amount,
                            credit_amount=credit_amount,
                            closing_balance=closing_balance,
                            amount=main_amount,  # Credit if available, otherwise debit
                            description=str(desc_val)[:500] if desc_val else '',
                            counterparty_account=str(counterparty_val)[:500] if counterparty_val else '',
                            payer_name='',  # Will be parsed from counterparty_account later if needed
                            status=BankTransaction.STATUS_PENDING
                        )
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f'Мөр {row_num}: {str(e)}')
                
                if imported_count > 0:
                    messages.success(request, f'✅ {imported_count} гүйлгээ амжилттай импортлогдлоо!')
                if skipped_count > 0:
                    messages.info(request, f'ℹ️ {skipped_count} гүйлгээ алгасагдсан (давхардсан эсвэл хоосон)')
                if imported_count == 0 and skipped_count == 0:
                    messages.warning(request, 'Импортлогдсон гүйлгээ байхгүй')
                
                if errors:
                    for error in errors[:10]:  # Show first 10 errors
                        messages.warning(request, error)
                    if len(errors) > 10:
                        messages.warning(request, f'... болон {len(errors) - 10} бусад алдаа')
                
                print(f"DEBUG: Total imported={imported_count}, skipped={skipped_count}, errors={len(errors)}")
                
                return redirect('bank_transaction_list')
                
            except Exception as e:
                messages.error(request, f'Файл уншихад алдаа гарлаа: {str(e)}')
    else:
        form = BankTransactionUploadForm()
    
    context = {
        'form': form,
    }
    return render(request, 'aikido_app/bank_transaction_upload.html', context)


@login_required
def delete_payment_allocation(request, allocation_id):
    """Төлбөрийн хуваарилалт устгах"""
    allocation = get_object_or_404(PaymentAllocation, id=allocation_id)
    transaction_id = allocation.bank_transaction.id
    
    if request.method == 'POST':
        allocation.delete()
        
        # Update transaction status
        transaction = BankTransaction.objects.get(id=transaction_id)
        transaction.update_status()
        
        messages.success(request, 'Төлбөрийн хуваарилалт устгагдлаа!')
    
    return redirect('bank_transaction_match', transaction_id=transaction_id)

@login_required
def delete_expense_allocation(request, allocation_id):
    """Зардлын хуваарилалт устгах"""
    allocation = get_object_or_404(ExpenseAllocation, id=allocation_id)
    transaction_id = allocation.bank_transaction.id
    
    if request.method == 'POST':
        allocation.delete()
        
        # Update transaction status
        transaction = BankTransaction.objects.get(id=transaction_id)
        transaction.update_status()
        
        messages.success(request, 'Зардлын хуваарилалт устгагдлаа!')
    
    return redirect('bank_transaction_match', transaction_id=transaction_id)

@login_required
def delete_instructor_payment_allocation(request, allocation_id):
    """Багшийн төлбөрийн хуваарилалт устгах"""
    allocation = get_object_or_404(InstructorPaymentAllocation, id=allocation_id)
    transaction_id = allocation.bank_transaction.id
    instructor_payment = allocation.instructor_payment
    
    if request.method == 'POST':
        allocation.delete()
        
        # Update instructor payment's paid_amount
        instructor_payment.update_paid_amount()
        
        # Update transaction status
        transaction = BankTransaction.objects.get(id=transaction_id)
        transaction.update_status()
        
        messages.success(request, 'Багшийн төлбөрийн хуваарилалт устгагдлаа!')
    
    return redirect('bank_transaction_match', transaction_id=transaction_id)

@login_required
def delete_seminar_payment_allocation(request, allocation_id):
    """Семинарын төлбөрийн хуваарилалт устгах"""
    allocation = get_object_or_404(SeminarPaymentAllocation, id=allocation_id)
    transaction_id = allocation.bank_transaction.id
    
    if request.method == 'POST':
        allocation.delete()
        
        # Update transaction status
        transaction = BankTransaction.objects.get(id=transaction_id)
        transaction.update_status()
        
        messages.success(request, 'Семинарын төлбөрийн хуваарилалт устгагдлаа!')
    
    return redirect('bank_transaction_match', transaction_id=transaction_id)

@login_required
def delete_membership_payment_allocation(request, allocation_id):
    """Гишүүнчлэлийн төлбөрийн хуваарилалт устгах"""
    allocation = get_object_or_404(MembershipPaymentAllocation, id=allocation_id)
    transaction_id = allocation.bank_transaction.id
    
    if request.method == 'POST':
        allocation.delete()
        
        # Update transaction status
        transaction = BankTransaction.objects.get(id=transaction_id)
        transaction.update_status()
        
        messages.success(request, 'Гишүүнчлэлийн төлбөрийн хуваарилалт устгагдлаа!')
    
    return redirect('bank_transaction_match', transaction_id=transaction_id)

@login_required
def delete_income_allocation(request, allocation_id):
    """Орлогын хуваарилалт устгах"""
    allocation = get_object_or_404(IncomeAllocation, id=allocation_id)
    
    if request.method == 'POST':
        transaction_id = allocation.bank_transaction.id
        allocation.delete()
        
        # Update transaction status
        transaction = BankTransaction.objects.get(id=transaction_id)
        transaction.update_status()
        
        messages.success(request, 'Орлогын хуваарилалт устгагдлаа!')
        return redirect('bank_transaction_match', transaction_id=transaction_id)
    
    return render(request, 'aikido_app/delete_income_allocation.html', {'allocation': allocation})

@login_required
def bank_transaction_match(request, transaction_id):
    """Банкны гүйлгээг сурагч эсвэл зардалтай холбох"""
    transaction = get_object_or_404(BankTransaction, id=transaction_id)
    
    # Determine if this is income (credit) or expense (debit)
    is_income = transaction.credit_amount and transaction.credit_amount > 0
    is_expense = transaction.debit_amount and transaction.debit_amount != 0
    
    if request.method == 'POST':
        if is_income:
            # Check income type from form
            income_type = request.POST.get('income_type', 'student_payment')
            
            # Debug logging
            print(f"DEBUG: income_type = {income_type}")
            print(f"DEBUG: POST data = {dict(request.POST)}")
            
            if income_type == 'student_payment':
                # Handle student payment allocations
                student_ids = request.POST.getlist('student_id[]')
                payment_months = request.POST.getlist('payment_month[]')
                amounts = request.POST.getlist('amount[]')
                notes_list = request.POST.getlist('notes[]')
                
                for i in range(len(student_ids)):
                    if student_ids[i] and payment_months[i] and amounts[i]:
                        try:
                            student = Student.objects.get(id=student_ids[i])
                            amount = Decimal(amounts[i])
                            payment_month = datetime.strptime(payment_months[i], '%Y-%m').date()
                            
                            instructor = None
                            if hasattr(request.user, 'instructor_profile'):
                                instructor = request.user.instructor_profile
                            
                            PaymentAllocation.objects.create(
                                bank_transaction=transaction,
                                student=student,
                                payment_month=payment_month,
                                amount=amount,
                                notes=notes_list[i] if i < len(notes_list) else '',
                                created_by=instructor
                            )
                        except Exception as e:
                            messages.error(request, f'Алдаа гарлаа: {str(e)}')
                            return redirect('bank_transaction_match', transaction_id=transaction_id)
                
                messages.success(request, f'{len([x for x in student_ids if x])} хуваарилалт амжилттай үүслээ!')
            
            elif income_type == 'other_income':
                # Handle other income allocations
                income_categories = request.POST.getlist('income_category[]')
                new_categories = request.POST.getlist('new_income_category[]')
                income_dates = request.POST.getlist('income_date[]')
                amounts = request.POST.getlist('amount[]')
                notes_list = request.POST.getlist('notes[]')
                
                # Use the maximum length to iterate through all rows
                max_len = max(len(income_categories), len(new_categories), len(income_dates), len(amounts))
                
                for i in range(max_len):
                    category_id = income_categories[i] if i < len(income_categories) else ''
                    new_category_name = new_categories[i] if i < len(new_categories) else ''
                    income_date_str = income_dates[i] if i < len(income_dates) else ''
                    amount_str = amounts[i] if i < len(amounts) else ''
                    
                    # Skip if no amount provided
                    if not amount_str:
                        continue
                    
                    # Must have either category_id or new_category_name
                    if not category_id and not new_category_name:
                        continue
                    
                    # Must have income_date
                    if not income_date_str:
                        continue
                        
                    try:
                        # Get or create category
                        if new_category_name:
                            category, _ = IncomeCategory.objects.get_or_create(name=new_category_name.strip())
                        elif category_id:
                            category = IncomeCategory.objects.get(id=category_id)
                        else:
                            continue
                        
                        amount = Decimal(amount_str)
                        income_date = datetime.strptime(income_date_str, '%Y-%m-%d').date()
                        
                        instructor = None
                        if hasattr(request.user, 'instructor_profile'):
                            instructor = request.user.instructor_profile
                        
                        IncomeAllocation.objects.create(
                            bank_transaction=transaction,
                            income_category=category,
                            income_date=income_date,
                            amount=amount,
                            notes=notes_list[i] if i < len(notes_list) else '',
                            created_by=instructor
                        )
                    except Exception as e:
                        messages.error(request, f'Алдаа гарлаа: {str(e)}')
                        return redirect('bank_transaction_match', transaction_id=transaction_id)
                
                messages.success(request, 'Орлогын хуваарилалт амжилттай үүслээ!')
            
            elif income_type == 'seminar_payment':
                # Handle seminar payment allocations
                student_ids = request.POST.getlist('seminar_student_id[]')
                seminar_ids = request.POST.getlist('seminar_id[]')
                amounts = request.POST.getlist('amount[]')
                notes_list = request.POST.getlist('notes[]')
                
                for i in range(len(student_ids)):
                    if student_ids[i] and seminar_ids[i] and i < len(amounts) and amounts[i]:
                        try:
                            student = Student.objects.get(id=student_ids[i])
                            seminar = Seminar.objects.get(id=seminar_ids[i])
                            amount = Decimal(amounts[i])
                            
                            instructor = None
                            if hasattr(request.user, 'instructor_profile'):
                                instructor = request.user.instructor_profile
                            
                            SeminarPaymentAllocation.objects.create(
                                bank_transaction=transaction,
                                student=student,
                                seminar=seminar,
                                amount=amount,
                                notes=notes_list[i] if i < len(notes_list) else '',
                                created_by=instructor
                            )
                        except Exception as e:
                            messages.error(request, f'Алдаа гарлаа: {str(e)}')
                            return redirect('bank_transaction_match', transaction_id=transaction_id)
                
                messages.success(request, 'Семинарын төлбөр амжилттай холбогдлоо!')
            
            elif income_type == 'membership_payment':
                # Handle membership payment allocations
                student_ids = request.POST.getlist('membership_student_id[]')
                payment_months = request.POST.getlist('membership_month[]')
                amounts = request.POST.getlist('amount[]')
                notes_list = request.POST.getlist('notes[]')
                
                for i in range(len(student_ids)):
                    if student_ids[i] and i < len(payment_months) and payment_months[i] and i < len(amounts) and amounts[i]:
                        try:
                            student = Student.objects.get(id=student_ids[i])
                            amount = Decimal(amounts[i])
                            payment_month = datetime.strptime(payment_months[i], '%Y-%m').date()
                            
                            instructor = None
                            if hasattr(request.user, 'instructor_profile'):
                                instructor = request.user.instructor_profile
                            
                            MembershipPaymentAllocation.objects.create(
                                bank_transaction=transaction,
                                student=student,
                                payment_month=payment_month,
                                amount=amount,
                                notes=notes_list[i] if i < len(notes_list) else '',
                                created_by=instructor
                            )
                        except Exception as e:
                            messages.error(request, f'Алдаа гарлаа: {str(e)}')
                            return redirect('bank_transaction_match', transaction_id=transaction_id)
                
                messages.success(request, 'Гишүүнчлэлийн төлбөр амжилттай холбогдлоо!')
            
            # Update transaction status
            transaction.update_status()
            
            # Check if there's still remaining amount
            remaining = transaction.get_remaining_amount()
            if remaining > 0:
                messages.info(request, f'Үлдсэн дүн: {remaining:,.0f}₮')
                return redirect('bank_transaction_match', transaction_id=transaction_id)
            else:
                messages.success(request, 'Гүйлгээ бүрэн хуваарилагдлаа!')
                return redirect('bank_transaction_list')
        
        elif is_expense:
            # Check if this is instructor/federation payment
            expense_type = request.POST.get('expense_type', 'regular')
            
            if expense_type == 'instructor_payment':
                # Handle instructor payment allocation (partial payments allowed)
                payment_ids = request.POST.getlist('instructor_payment_id[]')
                amounts = request.POST.getlist('instructor_payment_amount[]')
                notes_list = request.POST.getlist('instructor_payment_notes[]')
                
                instructor = None
                if hasattr(request.user, 'instructor_profile'):
                    instructor = request.user.instructor_profile
                
                for i, payment_id in enumerate(payment_ids):
                    if payment_id and i < len(amounts) and amounts[i]:
                        try:
                            payment = MonthlyInstructorPayment.objects.get(id=payment_id)
                            amount = Decimal(amounts[i])
                            
                            # Validate amount doesn't exceed remaining
                            remaining = payment.get_remaining_amount()
                            if amount > remaining:
                                messages.error(request, f'{payment.instructor}-ийн төлбөр: {amount}₮ нь үлдэгдэл {remaining}₮-с их байна!')
                                continue
                            
                            # Create allocation
                            InstructorPaymentAllocation.objects.create(
                                bank_transaction=transaction,
                                instructor_payment=payment,
                                amount=amount,
                                notes=notes_list[i] if i < len(notes_list) else '',
                                created_by=instructor
                            )
                            
                            # Update payment's paid_amount
                            payment.update_paid_amount()
                            
                        except MonthlyInstructorPayment.DoesNotExist:
                            messages.error(request, f'Төлбөр #{payment_id} олдсонгүй!')
                        except Exception as e:
                            messages.error(request, f'Алдаа гарлаа: {str(e)}')
                
                messages.success(request, 'Багшийн төлбөр амжилттай холбогдлоо!')
                
            elif expense_type == 'federation_payment':
                # Handle federation payment linking
                payment_ids = request.POST.getlist('federation_payment_id[]')
                
                for payment_id in payment_ids:
                    if payment_id:
                        try:
                            payment = MonthlyFederationPayment.objects.get(id=payment_id)
                            payment.bank_transaction = transaction
                            payment.is_paid = True
                            payment.paid_date = datetime.now().date()
                            payment.save()
                        except MonthlyFederationPayment.DoesNotExist:
                            messages.error(request, f'Төлбөр #{payment_id} олдсонгүй!')
                
                messages.success(request, 'Холбооны төлбөр амжилттай холбогдлоо!')
                
            else:
                # Handle regular expense allocations
                expense_categories = request.POST.getlist('expense_category[]')
                new_categories = request.POST.getlist('new_category[]')
                expense_dates = request.POST.getlist('expense_date[]')
                amounts = request.POST.getlist('amount[]')
                notes_list = request.POST.getlist('notes[]')
                
                for i in range(len(expense_categories)):
                    category_id = expense_categories[i] if i < len(expense_categories) else ''
                    new_category_name = new_categories[i] if i < len(new_categories) else ''
                    
                    if (category_id or new_category_name) and i < len(expense_dates) and i < len(amounts):
                        if not amounts[i]:
                            continue
                            
                        try:
                            # Get or create category
                            if new_category_name:
                                category, _ = ExpenseCategory.objects.get_or_create(name=new_category_name.strip())
                            elif category_id:
                                category = ExpenseCategory.objects.get(id=category_id)
                            else:
                                continue
                            
                            amount = Decimal(amounts[i])
                            expense_date = datetime.strptime(expense_dates[i], '%Y-%m-%d').date()
                            
                            instructor = None
                            if hasattr(request.user, 'instructor_profile'):
                                instructor = request.user.instructor_profile
                            
                            ExpenseAllocation.objects.create(
                                bank_transaction=transaction,
                                expense_category=category,
                                expense_date=expense_date,
                                amount=amount,
                                notes=notes_list[i] if i < len(notes_list) else '',
                                created_by=instructor
                            )
                        except Exception as e:
                            messages.error(request, f'Алдаа гарлаа: {str(e)}')
                            return redirect('bank_transaction_match', transaction_id=transaction_id)
                
                messages.success(request, 'Зардлын хуваарилалт амжилттай үүслээ!')
        
        # Update transaction status
        transaction.update_status()
        
        # Check if there's still remaining amount
        remaining = transaction.get_remaining_amount()
        if remaining > 0:
            messages.info(request, f'Үлдсэн дүн: {remaining:,.0f}₮')
            return redirect('bank_transaction_match', transaction_id=transaction_id)
        else:
            messages.success(request, 'Гүйлгээ бүрэн хуваарилагдлаа!')
            return redirect('bank_transaction_list')
    
    # GET request - show form
    context = {
        'transaction': transaction,
        'is_income': is_income,
        'is_expense': is_expense,
        'remaining_amount': transaction.get_remaining_amount(),
    }
    
    if is_income:
        students = Student.objects.filter(is_active=True).order_by('first_name', 'last_name')
        allocations = transaction.allocations.all()
        seminars = Seminar.objects.filter(is_active=True).order_by('-seminar_date')
        income_categories = IncomeCategory.objects.all().order_by('name')
        income_allocations = transaction.income_allocations.all()
        seminar_allocations = transaction.seminar_allocations.all()
        membership_allocations = transaction.membership_allocations.all()
        context.update({
            'students': students,
            'allocations': allocations,
            'seminars': seminars,
            'income_categories': income_categories,
            'income_allocations': income_allocations,
            'seminar_allocations': seminar_allocations,
            'membership_allocations': membership_allocations,
        })
    elif is_expense:
        expense_categories = ExpenseCategory.objects.all().order_by('name')
        expense_allocations = transaction.expense_allocations.all()
        instructor_payment_allocations = transaction.instructor_payment_allocations.all()
        
        # Get unpaid/partially paid instructor and federation payments
        unpaid_instructor_payments = MonthlyInstructorPayment.objects.filter(
            paid_amount__lt=F('instructor_share_amount')
        ).select_related('instructor', 'class_type').order_by('-month', 'instructor__last_name')
        
        unpaid_federation_payments = MonthlyFederationPayment.objects.filter(
            is_paid=False
        ).select_related('class_type').order_by('-month', 'class_type')
        
        context.update({
            'expense_categories': expense_categories,
            'expense_allocations': expense_allocations,
            'instructor_payment_allocations': instructor_payment_allocations,
            'unpaid_instructor_payments': unpaid_instructor_payments,
            'unpaid_federation_payments': unpaid_federation_payments,
        })
    
    return render(request, 'aikido_app/bank_transaction_match.html', context)


@login_required
def monthly_payment_report(request):
    """Сарын төлбөрийн тайлан - Багш болон холбооны хуваарилалт"""
    # Check if user is an instructor (non-staff)
    is_instructor_user = hasattr(request.user, 'instructor_profile') and not request.user.is_staff
    
    # Get filter parameters
    month_str = request.GET.get('month')
    year_str = request.GET.get('year')
    view_mode = request.GET.get('view_mode', 'month')  # month, year, all
    
    month_date = None
    year_filter = None
    
    if view_mode == 'month':
        # Month view
        if month_str:
            try:
                year, month_num = map(int, month_str.split('-'))
                month_date = datetime(year, month_num, 1).date()
            except (ValueError, AttributeError):
                month_date = datetime.now().date().replace(day=1)
        else:
            month_date = datetime.now().date().replace(day=1)
    elif view_mode == 'year':
        # Year view
        if year_str:
            try:
                year_filter = int(year_str)
            except (ValueError, TypeError):
                year_filter = datetime.now().year
        else:
            year_filter = datetime.now().year
    # else: view_mode == 'all' - no filters
    
    # Get all class types or filter by instructor's allowed class types
    if is_instructor_user:
        instructor_profile = request.user.instructor_profile
        allowed_class_types = instructor_profile.allowed_class_types.all()
        if allowed_class_types.exists():
            class_types = allowed_class_types
        else:
            class_types = ClassType.objects.all()
    else:
        class_types = ClassType.objects.all()
    
    # Build report data for each class type
    report_data = []
    for class_type in class_types:
        # Build query based on view mode
        payment_filter = {'student__class_types': class_type}
        
        if view_mode == 'month' and month_date:
            payment_filter['payment_month__year'] = month_date.year
            payment_filter['payment_month__month'] = month_date.month
        elif view_mode == 'year' and year_filter:
            payment_filter['payment_month__year'] = year_filter
        # else: view_mode == 'all' - no date filters
        
        # Get total payments collected for this class type
        payments = PaymentAllocation.objects.filter(**payment_filter)
        total_collected = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        student_count = payments.values('student').distinct().count()
        
        # Calculate splits
        federation_share = total_collected * Decimal('0.50')  # 50% for federation
        instructor_pool = total_collected * Decimal('0.50')   # 50% for instructors
        
        # Get federation payment record (if exists) - only for month view
        federation_payment = None
        if view_mode == 'month' and month_date:
            federation_payment = MonthlyFederationPayment.objects.filter(
                class_type=class_type,
                month=month_date
            ).first()
        
        # Get instructor payments - only for month view
        instructor_payments = MonthlyInstructorPayment.objects.none()
        if view_mode == 'month' and month_date:
            instructor_payments = MonthlyInstructorPayment.objects.filter(
                class_type=class_type,
                month=month_date
            ).select_related('instructor').order_by('role', 'instructor__last_name')
        
        # Calculate instructor totals
        instructor_total = instructor_payments.aggregate(total=Sum('instructor_share_amount'))['total'] or Decimal('0.00')
        
        # Group by role
        lead_payments = instructor_payments.filter(role=InstructorAssignment.LEAD)
        assistant_payments = instructor_payments.filter(role=InstructorAssignment.ASSISTANT)
        
        report_data.append({
            'class_type': class_type,
            'total_collected': total_collected,
            'federation_share': federation_share,
            'instructor_total': instructor_total,
            'instructor_pool': instructor_pool,
            'student_count': student_count,
            'lead_payments': lead_payments,
            'assistant_payments': assistant_payments,
            'federation_payment': federation_payment,
        })
    
    # Calculate grand totals
    grand_total_collected = Decimal('0.00')
    grand_federation_share = Decimal('0.00')
    grand_instructor_total = Decimal('0.00')
    
    for item in report_data:
        grand_total_collected += item['total_collected']
        grand_federation_share += item['federation_share']
        grand_instructor_total += item['instructor_pool']  # Use instructor_pool instead of instructor_total
    
    context = {
        'month_date': month_date,
        'year_filter': year_filter,
        'view_mode': view_mode,
        'report_data': report_data,
        'grand_total_collected': grand_total_collected,
        'grand_federation_share': grand_federation_share,
        'grand_instructor_total': grand_instructor_total,
        'available_years': range(2020, datetime.now().year + 2),  # 2020 to next year
    }
    
    return render(request, 'aikido_app/monthly_payment_report.html', context)


@login_required
def instructor_payment_list(request):
    """Багшийн төлбөрийн жагсаалт"""
    # Check if user is an instructor (non-staff)
    is_instructor_user = hasattr(request.user, 'instructor_profile') and not request.user.is_staff
    can_view_all = False
    
    # Filter options
    month_str = request.GET.get('month')
    instructor_id = request.GET.get('instructor')
    class_type_id = request.GET.get('class_type')
    is_paid = request.GET.get('is_paid')
    
    # Base query
    payments = MonthlyInstructorPayment.objects.select_related(
        'instructor', 'class_type'
    ).order_by('-month', 'class_type', 'instructor__last_name')
    
    # If instructor user, check permissions
    if is_instructor_user:
        instructor_profile = request.user.instructor_profile
        
        # Check if can view all payments
        if instructor_profile.can_view_all_payments:
            can_view_all = True
            # Can see all instructors but still respect class type restrictions
            allowed_class_types = instructor_profile.allowed_class_types.all()
            if allowed_class_types.exists():
                payments = payments.filter(class_type__in=allowed_class_types)
        else:
            # Only their own payments
            payments = payments.filter(instructor=instructor_profile)
            
            # Further filter by allowed class types if set
            allowed_class_types = instructor_profile.allowed_class_types.all()
            if allowed_class_types.exists():
                payments = payments.filter(class_type__in=allowed_class_types)
            
            # Don't allow them to filter by other instructors
            instructor_id = None
    
    # Apply filters
    if month_str:
        try:
            year, month = map(int, month_str.split('-'))
            month_date = datetime(year, month, 1).date()
            payments = payments.filter(month=month_date)
        except (ValueError, AttributeError):
            pass
    
    if instructor_id:
        payments = payments.filter(instructor_id=instructor_id)
    
    if class_type_id:
        payments = payments.filter(class_type_id=class_type_id)
    
    if is_paid == 'yes':
        payments = payments.filter(is_paid=True)
    elif is_paid == 'no':
        payments = payments.filter(is_paid=False)
    
    # Calculate summary
    total_amount = payments.aggregate(total=Sum('instructor_share_amount'))['total'] or Decimal('0.00')
    paid_amount = payments.filter(is_paid=True).aggregate(total=Sum('instructor_share_amount'))['total'] or Decimal('0.00')
    unpaid_amount = payments.filter(is_paid=False).aggregate(total=Sum('instructor_share_amount'))['total'] or Decimal('0.00')
    
    # Calculate detailed breakdown by class type and role for understanding
    calculation_details = []
    if month_str:
        try:
            year, month = map(int, month_str.split('-'))
            month_date = datetime(year, month, 1).date()
            
            for class_type in ClassType.objects.all():
                # Get payments for this class type in this month
                class_payments = PaymentAllocation.objects.filter(
                    payment_month=month_date,
                    student__class_types=class_type
                )
                total_collected = class_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                
                if total_collected > 0:
                    instructor_pool = total_collected * Decimal('0.50')  # 50% for instructors
                    lead_pool = instructor_pool * Decimal('0.60')  # 60% for leads
                    assistant_pool = instructor_pool * Decimal('0.40')  # 40% for assistants
                    
                    # Count total classes in this month
                    lead_classes = InstructorAssignment.objects.filter(
                        session__date__year=year,
                        session__date__month=month,
                        session__class_type=class_type,
                        role='LEAD'
                    ).count()
                    
                    assistant_classes = InstructorAssignment.objects.filter(
                        session__date__year=year,
                        session__date__month=month,
                        session__class_type=class_type,
                        role='ASSISTANT'
                    ).count()
                    
                    per_lead_class = (lead_pool / lead_classes) if lead_classes > 0 else Decimal('0.00')
                    per_assistant_class = (assistant_pool / assistant_classes) if assistant_classes > 0 else Decimal('0.00')
                    
                    calculation_details.append({
                        'class_type': class_type,
                        'total_collected': total_collected,
                        'instructor_pool': instructor_pool,
                        'lead_pool': lead_pool,
                        'assistant_pool': assistant_pool,
                        'lead_classes': lead_classes,
                        'assistant_classes': assistant_classes,
                        'per_lead_class': per_lead_class,
                        'per_assistant_class': per_assistant_class,
                    })
        except (ValueError, AttributeError):
            pass
    
    context = {
        'payments': payments,
        'instructors': Instructor.objects.filter(is_active=True).order_by('last_name'),
        'class_types': ClassType.objects.all(),
        'total_amount': total_amount,
        'paid_amount': paid_amount,
        'unpaid_amount': unpaid_amount,
        'calculation_details': calculation_details,
        'selected_month': month_str,
        'is_instructor_user': is_instructor_user,
        'can_view_all': can_view_all,
        'filters': {
            'month': month_str,
            'instructor': instructor_id,
            'class_type': class_type_id,
            'is_paid': is_paid,
        }
    }
    
    return render(request, 'aikido_app/instructor_payment_list.html', context)


@login_required
def mark_instructor_payment_paid(request, payment_id):
    """Багшийн төлбөрийг олгосон гэж тэмдэглэх"""
    payment = get_object_or_404(MonthlyInstructorPayment, id=payment_id)
    
    if request.method == 'POST':
        bank_transaction_id = request.POST.get('bank_transaction_id')
        
        # If bank transaction is selected, link it
        if bank_transaction_id:
            try:
                bank_transaction = BankTransaction.objects.get(id=bank_transaction_id)
                payment.bank_transaction = bank_transaction
            except BankTransaction.DoesNotExist:
                messages.error(request, 'Банкны гүйлгээ олдсонгүй!')
                return redirect('instructor_payment_list')
        
        payment.is_paid = True
        payment.paid_date = datetime.now().date()
        payment.save()
        messages.success(request, f'{payment.instructor} - {payment.instructor_share_amount}₮ олгосон гэж тэмдэглэгдлээ!')
        return redirect('instructor_payment_list')
    
    # GET - show form with available bank transactions (debit transactions for payment)
    # Get recent debit transactions that could be instructor payments
    recent_transactions = BankTransaction.objects.filter(
        debit_amount__gt=0
    ).order_by('-transaction_date')[:50]
    
    context = {
        'payment': payment,
        'recent_transactions': recent_transactions,
    }
    return render(request, 'aikido_app/confirm_payment.html', context)


@login_required
def federation_payment_list(request):
    """Холбооны төлбөрийн жагсаалт"""
    # Filter options
    month_str = request.GET.get('month')
    class_type_id = request.GET.get('class_type')
    is_paid = request.GET.get('is_paid')
    
    # Base query
    payments = MonthlyFederationPayment.objects.select_related(
        'class_type'
    ).order_by('-month', 'class_type')
    
    # Apply filters
    if month_str:
        try:
            year, month = map(int, month_str.split('-'))
            month_date = datetime(year, month, 1).date()
            payments = payments.filter(month=month_date)
        except (ValueError, AttributeError):
            pass
    
    if class_type_id:
        payments = payments.filter(class_type_id=class_type_id)
    
    if is_paid == 'yes':
        payments = payments.filter(is_paid=True)
    elif is_paid == 'no':
        payments = payments.filter(is_paid=False)
    
    # Calculate summary
    total_amount = payments.aggregate(total=Sum('federation_share_amount'))['total'] or Decimal('0.00')
    paid_amount = payments.filter(is_paid=True).aggregate(total=Sum('federation_share_amount'))['total'] or Decimal('0.00')
    unpaid_amount = payments.filter(is_paid=False).aggregate(total=Sum('federation_share_amount'))['total'] or Decimal('0.00')
    
    context = {
        'payments': payments,
        'class_types': ClassType.objects.all(),
        'total_amount': total_amount,
        'paid_amount': paid_amount,
        'unpaid_amount': unpaid_amount,
        'filters': {
            'month': month_str,
            'class_type': class_type_id,
            'is_paid': is_paid,
        }
    }
    
    return render(request, 'aikido_app/federation_payment_list.html', context)


@login_required
def mark_federation_payment_paid(request, payment_id):
    """Холбооны төлбөрийг олгосон гэж тэмдэглэх"""
    payment = get_object_or_404(MonthlyFederationPayment, id=payment_id)
    
    if request.method == 'POST':
        bank_transaction_id = request.POST.get('bank_transaction_id')
        
        # If bank transaction is selected, link it
        if bank_transaction_id:
            try:
                bank_transaction = BankTransaction.objects.get(id=bank_transaction_id)
                payment.bank_transaction = bank_transaction
            except BankTransaction.DoesNotExist:
                messages.error(request, 'Банкны гүйлгээ олдсонгүй!')
                return redirect('federation_payment_list')
        
        payment.is_paid = True
        payment.paid_date = datetime.now().date()
        payment.save()
        messages.success(request, f'{payment.class_type} - {payment.federation_share_amount}₮ олгосон гэж тэмдэглэгдлээ!')
        return redirect('federation_payment_list')
    
    # GET - show form with available bank transactions (debit transactions for payment)
    recent_transactions = BankTransaction.objects.filter(
        debit_amount__gt=0
    ).order_by('-transaction_date')[:50]
    
    context = {
        'payment': payment,
        'recent_transactions': recent_transactions,
    }
    return render(request, 'aikido_app/confirm_payment.html', context)


@login_required
def calculate_instructor_payments_from_attendance(request):
    """Ирцийн хуудаснаас багшийн төлбөр тооцоолох"""
    if request.method == 'POST':
        month_str = request.POST.get('month')
        
        if not month_str:
            messages.error(request, 'Сар сонгоно уу!')
            return redirect('attendance_record')
        
        try:
            # Parse month
            year, month = map(int, month_str.split('-'))
            month_date = datetime(year, month, 1).date()
            
            # Call management command logic
            from django.core.management import call_command
            from io import StringIO
            
            out = StringIO()
            call_command('calculate_monthly_payments', 
                        month=month_str,
                        recalculate=True,
                        stdout=out)
            
            output = out.getvalue()
            
            messages.success(request, f'{month_str} сарын багшийн төлбөр амжилттай тооцоологдлоо!')
            
            # Redirect back to attendance record page with the same month
            return redirect(f'/attendance/record/?month={month_str}')
            
        except Exception as e:
            messages.error(request, f'Алдаа гарлаа: {str(e)}')
            return redirect('attendance_record')
    
    return redirect('attendance_record')


        
    
